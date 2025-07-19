"""
改进的文件处理模块
"""
import csv
import json
import logging
import os
from pathlib import Path
from typing import Optional, List, Dict, Any, Union
from concurrent.futures import ThreadPoolExecutor, as_completed
import mimetypes
from tqdm import tqdm

from PyPDF2 import PdfReader
import openpyxl
from config import config

# 配置日志
logging.basicConfig(
    level=getattr(logging, config.get('logging.level', 'INFO')),
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(config.get('logging.file', 'app.log'), encoding='utf-8'),
        logging.StreamHandler()
    ]
)

logger = logging.getLogger(__name__)

class FileHandler:
    """文件处理类"""
    
    def __init__(self):
        self.supported_formats = set(config.get('processing.supported_formats', []))
        self.max_file_size = config.get('processing.max_file_size_mb', 100) * 1024 * 1024
    
    def validate_file(self, file_path: Union[str, Path]) -> bool:
        """验证文件是否有效"""
        file_path = Path(file_path)
        
        if not file_path.exists():
            logger.error(f"文件不存在: {file_path}")
            return False
        
        if not file_path.is_file():
            logger.error(f"路径不是文件: {file_path}")
            return False
        
        if file_path.stat().st_size > self.max_file_size:
            logger.error(f"文件太大: {file_path} ({file_path.stat().st_size} bytes)")
            return False
        
        if file_path.suffix.lower() not in self.supported_formats:
            logger.warning(f"不支持的文件格式: {file_path.suffix}")
            return False
        
        return True
    
    def read_file(self, file_path: Union[str, Path]) -> Optional[str]:
        """通用文件读取方法"""
        file_path = Path(file_path)
        
        if not self.validate_file(file_path):
            return None
        
        try:
            suffix = file_path.suffix.lower()
            
            if suffix == '.txt':
                return self._read_text_file(file_path)
            elif suffix == '.csv':
                return self._read_csv_file(file_path)
            elif suffix == '.json':
                return self._read_json_file(file_path)
            elif suffix == '.pdf':
                return self._read_pdf_file(file_path)
            elif suffix in ['.xlsx', '.xls']:
                return self._read_excel_file(file_path)
            else:
                # 尝试作为文本文件读取
                return self._read_text_file(file_path)
                
        except Exception as e:
            logger.error(f"读取文件失败 {file_path}: {e}")
            return None
    
    def _read_text_file(self, file_path: Path) -> Optional[str]:
        """读取文本文件"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        except UnicodeDecodeError:
            # 尝试其他编码
            for encoding in ['gbk', 'iso-8859-1', 'cp1252']:
                try:
                    with open(file_path, 'r', encoding=encoding) as f:
                        content = f.read()
                        logger.info(f"使用 {encoding} 编码读取文件: {file_path}")
                        return content
                except UnicodeDecodeError:
                    continue
            logger.error(f"无法确定文件编码: {file_path}")
            return None
    
    def _read_csv_file(self, file_path: Path) -> Optional[str]:
        """读取CSV文件"""
        try:
            rows = []
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                for row in reader:
                    rows.append(' '.join(str(cell) for cell in row))
            return '\n'.join(rows)
        except Exception as e:
            logger.error(f"读取CSV文件失败 {file_path}: {e}")
            return None
    
    def _read_json_file(self, file_path: Path) -> Optional[str]:
        """读取JSON文件"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return json.dumps(data, ensure_ascii=False, indent=2)
        except json.JSONDecodeError as e:
            logger.error(f"JSON格式错误 {file_path}: {e}")
            return None
    
    def _read_pdf_file(self, file_path: Path) -> Optional[str]:
        """读取PDF文件"""
        try:
            reader = PdfReader(str(file_path))
            text_parts = []
            for page in reader.pages:
                text = page.extract_text()
                if text.strip():
                    text_parts.append(text)
            return '\n'.join(text_parts)
        except Exception as e:
            logger.error(f"读取PDF文件失败 {file_path}: {e}")
            return None
    
    def _read_excel_file(self, file_path: Path) -> Optional[str]:
        """读取Excel文件"""
        try:
            workbook = openpyxl.load_workbook(str(file_path), read_only=True)
            text_parts = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_text = []
                for row in sheet.iter_rows(values_only=True):
                    row_text = ' '.join(str(cell) for cell in row if cell is not None)
                    if row_text.strip():
                        sheet_text.append(row_text)
                
                if sheet_text:
                    text_parts.append(f"Sheet: {sheet_name}\n" + '\n'.join(sheet_text))
            
            workbook.close()
            return '\n\n'.join(text_parts)
        except Exception as e:
            logger.error(f"读取Excel文件失败 {file_path}: {e}")
            return None
    
    def write_file(self, file_path: Union[str, Path], content: str, mode: str = 'w') -> bool:
        """写入文件"""
        file_path = Path(file_path)
        
        try:
            # 确保目录存在
            file_path.parent.mkdir(parents=True, exist_ok=True)
            
            with open(file_path, mode, encoding='utf-8') as f:
                f.write(content)
            
            logger.info(f"文件写入成功: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"写入文件失败 {file_path}: {e}")
            return False
    
    def batch_process(self, input_folder: Union[str, Path], 
                     output_folder: Union[str, Path],
                     processor_func) -> Dict[str, Any]:
        """批量处理文件"""
        input_folder = Path(input_folder)
        output_folder = Path(output_folder)
        
        if not input_folder.exists():
            logger.error(f"输入文件夹不存在: {input_folder}")
            return {"success": False, "error": "输入文件夹不存在"}
        
        # 创建输出文件夹
        output_folder.mkdir(parents=True, exist_ok=True)
        
        # 获取所有支持的文件
        files_to_process = []
        for file_path in input_folder.rglob('*'):
            if file_path.is_file() and self.validate_file(file_path):
                files_to_process.append(file_path)
        
        if not files_to_process:
            logger.warning("没有找到可处理的文件")
            return {"success": True, "processed": 0, "errors": 0}
        
        # 并行处理文件
        max_workers = config.get('processing.max_workers', 4)
        processed_count = 0
        error_count = 0
        
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # 提交任务
            future_to_file = {}
            for file_path in files_to_process:
                relative_path = file_path.relative_to(input_folder)
                output_path = output_folder / f"{relative_path.stem}.processed{relative_path.suffix}"
                
                future = executor.submit(self._process_single_file, 
                                       file_path, output_path, processor_func)
                future_to_file[future] = file_path
            
            # 收集结果
            with tqdm(total=len(files_to_process), desc="处理文件") as pbar:
                for future in as_completed(future_to_file):
                    file_path = future_to_file[future]
                    try:
                        success = future.result()
                        if success:
                            processed_count += 1
                        else:
                            error_count += 1
                    except Exception as e:
                        logger.error(f"处理文件时发生错误 {file_path}: {e}")
                        error_count += 1
                    
                    pbar.update(1)
        
        logger.info(f"批量处理完成: {processed_count} 成功, {error_count} 失败")
        return {
            "success": True,
            "processed": processed_count,
            "errors": error_count,
            "total": len(files_to_process)
        }
    
    def _process_single_file(self, input_path: Path, output_path: Path, 
                           processor_func) -> bool:
        """处理单个文件"""
        try:
            content = self.read_file(input_path)
            if content is None:
                return False
            
            processed_content = processor_func(content)
            if processed_content is None:
                return False
            
            return self.write_file(output_path, processed_content)
            
        except Exception as e:
            logger.error(f"处理单个文件失败 {input_path}: {e}")
            return False

# 全局文件处理器实例
file_handler = FileHandler()